import json
from openpyxl import load_workbook

path = r"G:\マイドライブ\Antigravity\xROAD_DataSearcher\点検実施工程表.xlsx"
book = load_workbook(path, read_only=False, data_only=True)
sheet = book.worksheets[0]

def color(cell):
    c = cell.fill.fgColor
    return c.rgb if c.type == "rgb" else f"{c.type}:{c.theme}:{c.tint}"

for r in range(2, 4):
    cells = []
    for c in range(1, 60):
        cell = sheet.cell(r, c)
        cinfo = color(cell)
        if cell.fill.fill_type == "solid" and cinfo not in ("00000000", "FFFFFFFF", "FF000000", "theme:0:-0.499984740745262"):
            cells.append({"cell": cell.coordinate, "value": cell.value, "style": cell.style_id, "color": cinfo})
    print(json.dumps({"row": r, "cells": cells}, ensure_ascii=True, default=str))

for source in ("properties", "defined_names"):
    if source == "properties":
        props = book.properties
        value = {"created": str(props.created), "modified": str(props.modified), "title": props.title, "subject": props.subject, "description": props.description}
    else:
        value = list(book.defined_names)
    print(json.dumps({source: value}, ensure_ascii=True, default=str))
